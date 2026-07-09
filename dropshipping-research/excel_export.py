"""Builds results.xlsx with TikTok/Instagram/YouTube/Shopify/AliExpress tabs + a Summary dashboard."""

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True)
TOP_SCORE_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")


def _autofit_and_style(worksheet, df: pd.DataFrame):
    worksheet.freeze_panes = "A2"

    for col_idx, column in enumerate(df.columns, start=1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")

        max_len = max(
            [len(str(column))] + [len(str(v)) for v in df[column].astype(str).tolist()[:500]]
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 60)

    if "score" in df.columns and not df.empty:
        score_col_idx = list(df.columns).index("score") + 1
        threshold = df["score"].quantile(0.9) if len(df) > 1 else 0
        for row_idx, value in enumerate(df["score"].tolist(), start=2):
            if pd.notna(value) and value >= threshold:
                worksheet.cell(row=row_idx, column=score_col_idx).fill = TOP_SCORE_FILL


def _build_summary(sheets: dict) -> pd.DataFrame:
    rows = []
    for sheet_name, df in sheets.items():
        if sheet_name == "Summary" or df.empty:
            continue

        group_col = "niche" if "niche" in df.columns else None
        if group_col is None:
            continue

        for niche, group in df.groupby(group_col):
            row = {
                "niche": niche,
                "platform": sheet_name,
                "records_found": len(group),
            }
            if "score" in group.columns and group["score"].notna().any():
                row["avg_score"] = round(group["score"].mean(), 2)
                top = group.loc[group["score"].idxmax()]
                row["top_score"] = top["score"]
                row["top_url"] = top.get("url", "")
            elif "mentions" in group.columns:
                top = group.loc[group["mentions"].idxmax()]
                row["avg_score"] = ""
                row["top_score"] = top["mentions"]
                row["top_url"] = top.get("url", "")
            else:
                row["avg_score"] = ""
                row["top_score"] = ""
                row["top_url"] = ""
            rows.append(row)

    summary_df = pd.DataFrame(rows, columns=[
        "niche", "platform", "records_found", "avg_score", "top_score", "top_url",
    ])
    return summary_df.sort_values(["niche", "platform"]).reset_index(drop=True)


def build_workbook(
    output_path: str,
    tiktok_df: pd.DataFrame,
    instagram_df: pd.DataFrame,
    youtube_df: pd.DataFrame,
    shopify_df: pd.DataFrame,
    aliexpress_df: pd.DataFrame,
):
    sheets = {
        "TikTok_Winners": tiktok_df,
        "Instagram_Winners": instagram_df,
        "YouTube_Winners": youtube_df,
        "Shopify_Competitors": shopify_df,
        "AliExpress_Winners": aliexpress_df,
    }
    sheets["Summary"] = _build_summary(sheets)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            _autofit_and_style(writer.sheets[sheet_name], df)

    return output_path
